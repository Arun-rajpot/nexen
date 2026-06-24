import Unclassified_share_capital
import Preference_share_capital
import Equity_share_capital
import Break_up_of_paid_up_share_capital
import Break_up_of_paid_up_share_capital_Mgt7A
import company_info
import Number_of_Registrar_and_Transfer_Agent
import Non_convertible_debentures
import Basic_Finacials
import Summary_of_Indebtedness
import Partly_convertible_debentures
import Fully_convertible_debentures
import A_Promoters
import Public_Other_than_promoters
import Board_of_Directors
import NUMBER_OF_PROMOTERS_MEMBERS_DEBENTURE_HOLDERS
import FinancialSummary
import pandas as pd

def update_company_info(company_information, BF):

    """
    Updates the Basic Financials dictionary with company information.
    Handles missing/empty values gracefully.
    """
    def safe_get(key, default=None):
        val = company_information.get(key, default)
        return val if val not in (None, '', 'NA', 'N/A') else default

    # Core fields
    BF['CIN'] = safe_get('cin')

    # Extract Year from Financial Year End Date
    fy_end = safe_get('Financial Year End Date')
    if fy_end and len(fy_end) >= 4:
        BF['Year'] = fy_end.split('/')[-1]
    else:
        BF['Year'] = None

    BF['Pan of Company'] = "" # safe_get('pan')
    BF['Company Name'] = safe_get('name').title()
    print(BF['Company Name'])
    BF['Website'] = safe_get('website')
    BF['Telephone'] = "" #safe_get('phone')
    BF['Financial Year Start Date'] = safe_get('Financial Year Start Date')
    BF['Financial Year End Date'] = fy_end

    # Share Capital Flag
    category = safe_get('category', '').lower()
    BF['If company has Share Capital'] = 'Yes' if 'company limited by shares' in category else 'No'

    # RTA fields
    BF['CIN of Registrar/Transfer Agent'] = None
    BF['Name of Registrar/Transfer Agent'] = None
    BF['Address of Registrar/Transfer Agent'] = None

    return BF


def update_break_up_of_paid_up_share_capital(break_up_data, BF):
    # print(break_up_data)

    def find_section(lst, key, n=1):
        count = 0
        for item in lst:
            if isinstance(item, dict):
                for k, v in item.items():
                    # fuzzy match: if key is prefix of k
                    if k.startswith(key):
                        count += 1
                        if count == n:
                            return v
        return {}

    equity_list = break_up_data.get('(i) Equity shares', [])

    # ===== Equity Main Sections =====
    boy_sec = find_section(equity_list, 'At the beginning of the year')
    inc_sec = find_section(equity_list, 'Increase during the year')
    dec_sec = find_section(equity_list, 'Decrease during the year')
    eoy_sec = find_section(equity_list, 'At the end of the year')

    # ===== Equity Increase Specific =====
    pi_sec   = find_section(equity_list, 'i Public Issues')
    ri_sec   = find_section(equity_list, 'ii Rights issue') or find_section(equity_list, 'i Rights issue')
    bi_sec   = find_section(equity_list, 'iii Bonus issue') or find_section(equity_list, 'ii Bonus issue')
    pp_sec   = find_section(equity_list, 'iv Private Placement/ Preferential\nallotment') or find_section(equity_list, 'iii Private Placement/ Preferential\nallotment')
    esop_sec = find_section(equity_list, 'v ESOPs') or find_section(equity_list, 'iv ESOPs')
    se_sec   = find_section(equity_list, 'vi Sweat equity shares allotted') or find_section(equity_list, 'v Sweat equity shares allotted')
    cps_sec  = find_section(equity_list, 'vii Conversion of Preference share') or find_section(equity_list, 'vi Conversion of Preference share')
    cd_sec   = find_section(equity_list, 'viii Conversion of Debentures') or find_section(equity_list, 'vii Conversion of Debentures')
    dr_sec   = find_section(equity_list, 'ix GDRs/ADRs')
    # oth_inc_sec = find_section(equity_list, 'x Others, specify\nDematerialisation of shares') or find_section(equity_list, "x Others, specify\nDemat")
    # print("===oth_inc_sec=",oth_inc_sec)

    oth_inc_sec = find_section(equity_list, "x Others, specify") or find_section(equity_list, "viii Others, specify")
    oth_dec_sec = find_section(equity_list, "iv Others, specify")
    # ===== Equity Decrease Specific =====
    bb_sec  = find_section(equity_list, 'i Buy-back of shares')
    sf_sec  = find_section(equity_list, 'ii Shares forfeited')
    rsc_sec = find_section(equity_list, 'iii Reduction of share capital')
    # oth_dec_sec = find_section(equity_list, 'iv Others, specify\nDematerialisation of shares') or find_section(equity_list, 'iv Others, specify\nDemat')

    # ===== Fill Equity Data =====
    BF['Break Up PC - No. of Shares at BoY'] = boy_sec.get('Total')
    BF['Break Up PC - Total Nominal Amount at BoY'] = boy_sec.get('Total Nominal Amount')
    BF['Break Up PC - Total PC Amount at BoY'] = boy_sec.get('Total Paid-up amount')
    BF['Break Up PC - Total Premium at BoY'] = boy_sec.get('Total premium')

    BF['Break Up PC - Inc No. of Shares at DY'] = inc_sec.get('Total')
    BF['Break Up PC - Inc Total Nominal Amount at DY'] = inc_sec.get('Total Nominal Amount')
    BF['Break Up PC - Inc Total PC Amount at DY'] = inc_sec.get('Total Paid-up amount')
    BF['Break Up PC - Inc Total Premium at DY'] = inc_sec.get('Total premium')

    # === Increase Specific Equity Share Mapping ===

    # Public Issue
    BF['Break Up PC - PI No. of Shares at DY'] = pi_sec.get('Total')
    BF['Break Up PC - PI Total Nominal Amount at DY'] = pi_sec.get('Total Nominal Amount')
    BF['Break Up PC - PI Total PC Amount at DY'] = pi_sec.get('Total Paid-up amount')
    BF['Break Up PC - PI Total Premium at DY'] = pi_sec.get('Total premium')

    # Rights Issue
    BF['Break Up PC - RI No. of Shares at DY'] = ri_sec.get('Total')
    BF['Break Up PC - RI Total Nominal Amount at DY'] = ri_sec.get('Total Nominal Amount')
    BF['Break Up PC - RI Total PC Amount at DY'] = ri_sec.get('Total Paid-up amount')
    BF['Break Up PC - RI Total Premium at DY'] = ri_sec.get('Total premium')

    # Bonus Issue
    BF['Break Up PC - BI No. of Shares at DY'] = bi_sec.get('Total')
    BF['Break Up PC - BI Total Nominal Amount at DY'] = bi_sec.get('Total Nominal Amount')
    BF['Break Up PC - BI Total PC Amount at DY'] = bi_sec.get('Total Paid-up amount')
    BF['Break Up PC - BI Total Premium at DY'] = bi_sec.get('Total premium')

    # Private Placement / Preferential
    BF['Break Up PC - PP No. of Shares at DY'] = pp_sec.get('Total')
    BF['Break Up PC - PP Total Nominal Amount at DY'] = pp_sec.get('Total Nominal Amount')
    BF['Break Up PC - PP Total PC Amount at DY'] = pp_sec.get('Total Paid-up amount')
    BF['Break Up PC - PP Total Premium at DY'] = pp_sec.get('Total premium')

    # ESOP
    BF['Break Up PC - ESOP No. of Shares at DY'] = esop_sec.get('Total')
    BF['Break Up PC - ESOP Total Nominal Amount at DY'] = esop_sec.get('Total Nominal Amount')
    BF['Break Up PC - ESOP Total PC Amount at DY'] = esop_sec.get('Total Paid-up amount')
    BF['Break Up PC - ESOP Total Premium at DY'] = esop_sec.get('Total premium')

    # Sweat Equity
    BF['Break Up PC - SE No. of Shares at DY'] = se_sec.get('Total')
    BF['Break Up PC - SE Total Nominal Amount at DY'] = se_sec.get('Total Nominal Amount')
    BF['Break Up PC - SE Total PC Amount at DY'] = se_sec.get('Total Paid-up amount')
    BF['Break Up PC - SE Total Premium at DY'] = se_sec.get('Total premium')

    # Conversion of Preference Shares
    BF['Break Up PC - CPS No. of Shares at DY'] = cps_sec.get('Total')
    BF['Break Up PC - CPS Total Nominal Amount at DY'] = cps_sec.get('Total Nominal Amount')
    BF['Break Up PC - CPS Total PC Amount at DY'] = cps_sec.get('Total Paid-up amount')
    BF['Break Up PC - CPS Total Premium at DY'] = cps_sec.get('Total premium')

    # Conversion of Debentures
    BF['Break Up PC - CD No. of Shares at DY'] = cd_sec.get('Total')
    BF['Break Up PC - CD Total Nominal Amount at DY'] = cd_sec.get('Total Nominal Amount')
    BF['Break Up PC - CD Total PC Amount at DY'] = cd_sec.get('Total Paid-up amount')
    BF['Break Up PC - CD Total Premium at DY'] = cd_sec.get('Total premium')

    # GDRs/ADRs
    BF['Break Up PC - DR No. of Shares at DY'] = dr_sec.get('Total')
    BF['Break Up PC - DR Total Nominal Amount at DY'] = dr_sec.get('Total Nominal Amount')
    BF['Break Up PC - DR Total PC Amount at DY'] = dr_sec.get('Total Paid-up amount')
    BF['Break Up PC - DR Total Premium at DY'] = dr_sec.get('Total premium')

    # Others
    BF['Break Up PC - OTH No. of Shares at DY'] = oth_inc_sec.get('Total')
    BF['Break Up PC - OTH Total Nominal Amount at DY'] = oth_inc_sec.get('Total Nominal Amount')
    BF['Break Up PC - OTH Total PC Amount at DY'] = oth_inc_sec.get('Total Paid-up amount')
    BF['Break Up PC - OTH Total Premium at DY'] = oth_inc_sec.get('Total premium')

    BF['Break Up PC - Dec. No. of Shares at DY'] = dec_sec.get('Total')
    BF['Break Up PC - Dec. Total Nominal Amount at DY'] = dec_sec.get('Total Nominal Amount')
    BF['Break Up PC - Dec. Total PC Amount at DY'] = dec_sec.get('Total Paid-up amount')
    BF['Break Up PC - Dec. Total Premium at DY'] = dec_sec.get('Total premium')

    # Buy-back of shares
    BF['Break Up PC - BB No. of Shares at DY'] = bb_sec.get('Total')
    BF['Break Up PC - BB Total Nominal Amount at DY'] = bb_sec.get('Total Nominal Amount')
    BF['Break Up PC - BB Total PC Amount at DY'] = bb_sec.get('Total Paid-up amount')
    BF['Break Up PC - BB Total Premium at DY'] = bb_sec.get('Total premium')

    # Shares forfeited
    BF['Break Up PC - SF No. of Shares at DY'] = sf_sec.get('Total')
    BF['Break Up PC - SF Total Nominal Amount at DY'] = sf_sec.get('Total Nominal Amount')
    BF['Break Up PC - SF Total PC Amount at DY'] = sf_sec.get('Total Paid-up amount')
    BF['Break Up PC - SF Total Premium at DY'] = sf_sec.get('Total premium')

    # Reduction of share capital
    BF['Break Up PC - RSC No. of Shares at DY'] = rsc_sec.get('Total')
    BF['Break Up PC - RSC Total Nominal Amount at DY'] = rsc_sec.get('Total Nominal Amount')
    BF['Break Up PC - RSC Total PC Amount at DY'] = rsc_sec.get('Total Paid-up amount')
    BF['Break Up PC - RSC Total Premium at DY'] = rsc_sec.get('Total premium')

    # Other decrease (Dematerialisation etc.)
    BF['Break Up PC - Dec. Oth No. of Shares at DY'] = oth_dec_sec.get('Total')
    BF['Break Up PC - Dec. Oth Total Nominal Amount at DY'] = oth_dec_sec.get('Total Nominal Amount')
    BF['Break Up PC - Dec. Oth Total PC Amount at DY'] = oth_dec_sec.get('Total Paid-up amount')
    BF['Break Up PC - Dec. Oth Total Premium at DY'] = oth_dec_sec.get('Total premium')

    BF['Break Up PC - No. of Shares at EoY'] = eoy_sec.get('Total')
    BF['Break Up PC - Total Nominal Amount at EoY'] = eoy_sec.get('Total Nominal Amount')
    BF['Break Up PC - Total PC Amount at EoY'] = eoy_sec.get('Total Paid-up amount')
    BF['Break Up PC - Total Premium at EoY'] = eoy_sec.get('Total premium')

    # ===== Preference Shares (same list!) =====
    p_boy_sec = find_section(equity_list, '(ii) Preference shares')
    p_boy_sec = p_boy_sec if p_boy_sec.get('Total') is not None else {}

    p_boystart = find_section(equity_list, 'At the beginning of the year', 2)
    p_incsec   = find_section(equity_list, 'Increase during the year',2)
    p_decsec   = find_section(equity_list, 'Decrease during the year',2)
    p_eoysec   = find_section(equity_list, 'At the end of the year',2)

    p_is_sec = find_section(equity_list, 'i Issues of shares')
    p_rs_sec = find_section(equity_list, 'ii Re-issue of forfeited shares')
    p_oth_inc_sec = find_section(equity_list, 'iii Others, specify\nNA') or find_section(equity_list, 'iii Others, specify')

    p_rd_sec = find_section(equity_list, 'i Redemption of shares')
    p_sf_sec = find_section(equity_list, 'ii Shares forfeited',2)
    p_rsc_sec = find_section(equity_list, 'iii Reduction of share capital',2)
    p_oth_dec_sec = find_section(equity_list, 'iv Others, specify\nNA',2) or find_section(equity_list, 'iv Others, specify',2) or find_section(equity_list, 'iii Others, specify',2)

    # ===== Fill Preference Data =====

    # --- Opening Balance (BoY) ---
    BF['Break Up PC - No. of Pref Shares at BoY'] = p_boystart.get('Total')
    BF['Break Up PC - Total Pref Nominal Amount at BoY'] = p_boystart.get('Total Nominal Amount')
    BF['Break Up PC - Total Pref PC Amount at BoY'] = p_boystart.get('Total Paid-up amount')
    BF['Break Up PC - Total Pref Premium at BoY'] = p_boystart.get('Total premium')

    # --- Total Increase ---
    BF['Break Up PC - Inc No. of Pref Shares at DY'] = p_incsec.get('Total')
    BF['Break Up PC - Inc Total Pref Nominal Amount at DY'] = p_incsec.get('Total Nominal Amount')
    BF['Break Up PC - Inc Total Pref PC Amount at DY'] = p_incsec.get('Total Paid-up amount')
    BF['Break Up PC - Inc Total Pref Premium at DY'] = p_incsec.get('Total premium')

    # --- Increase Specific ---
    BF['Break Up PC - IS No. of Pref Shares at DY'] = p_is_sec.get('Total')
    BF['Break Up PC - IS Total Pref Nominal Amount at DY'] = p_is_sec.get('Total Nominal Amount')
    BF['Break Up PC - IS Total Pref PC Amount at DY'] = p_is_sec.get('Total Paid-up amount')
    BF['Break Up PC - IS Total Pref Premium at DY'] = p_is_sec.get('Total premium')

    BF['Break Up PC - RS No. of Pref Shares at DY'] = p_rs_sec.get('Total')
    BF['Break Up PC - RS Total Pref Nominal Amount at DY'] = p_rs_sec.get('Total Nominal Amount')
    BF['Break Up PC - RS Total Pref PC Amount at DY'] = p_rs_sec.get('Total Paid-up amount')
    BF['Break Up PC - RS Total Pref Premium at DY'] = p_rs_sec.get('Total premium')

    BF['Break Up PC - Oth No. of Pref Shares at DY'] = p_oth_inc_sec.get('Total')
    BF['Break Up PC - Oth Total Pref Nominal Amount at DY'] = p_oth_inc_sec.get('Total Nominal Amount')
    BF['Break Up PC - Oth Total Pref PC Amount at DY'] = p_oth_inc_sec.get('Total Paid-up amount')
    BF['Break Up PC - Oth Total Pref Premium at DY'] = p_oth_inc_sec.get('Total premium')

    # --- Total Decrease ---
    BF['Break Up PC - Dec. No. of Pref Shares at DY'] = p_decsec.get('Total')
    BF['Break Up PC - Dec. Total Pref Nominal Amount at DY'] = p_decsec.get('Total Nominal Amount')
    BF['Break Up PC - Dec. Total Pref PC Amount at DY'] = p_decsec.get('Total Paid-up amount')
    BF['Break Up PC - Dec. Total Pref Premium at DY'] = p_decsec.get('Total premium')

    # --- Decrease Specific ---
    BF['Break Up PC - RD No. of Pref Shares at DY'] = p_rd_sec.get('Total')
    BF['Break Up PC - RD Total Pref Nominal Amount at DY'] = p_rd_sec.get('Total Nominal Amount')
    BF['Break Up PC - RD Total Pref PC Amount at DY'] = p_rd_sec.get('Total Paid-up amount')
    BF['Break Up PC - RD Total Pref Premium at DY'] = p_rd_sec.get('Total premium')

    BF['Break Up PC - SF No. of Pref Shares at DY'] = p_sf_sec.get('Total')
    BF['Break Up PC - SF Total Pref Nominal Amount at DY'] = p_sf_sec.get('Total Nominal Amount')
    BF['Break Up PC - SF Total Pref PC Amount at DY'] = p_sf_sec.get('Total Paid-up amount')
    BF['Break Up PC - SF Total Pref Premium at DY'] = p_sf_sec.get('Total premium')

    BF['Break Up PC - RSC No. of Pref Shares at DY'] = p_rsc_sec.get('Total')
    BF['Break Up PC - RSC Total Pref Nominal Amount at DY'] = p_rsc_sec.get('Total Nominal Amount')
    BF['Break Up PC - RSC Total Pref PC Amount at DY'] = p_rsc_sec.get('Total Paid-up amount')
    BF['Break Up PC - RSC Total Pref Premium at DY'] = p_rsc_sec.get('Total premium')

    BF['Break Up PC - Dec. Oth No. of Pref Shares at DY'] = p_oth_dec_sec.get('Total')
    BF['Break Up PC - Dec. Oth Total Pref Nominal Amount at DY'] = p_oth_dec_sec.get('Total Nominal Amount')
    BF['Break Up PC - Dec. Oth Total Pref PC Amount at DY'] = p_oth_dec_sec.get('Total Paid-up amount')
    BF['Break Up PC - Dec. Oth Total Pref Premium at DY'] = p_oth_dec_sec.get('Total premium')

    # --- Closing Balance (EoY) ---
    BF['Break Up PC - No. of Pref Shares at EoY'] = p_eoysec.get('Total')
    BF['Break Up PC - Total Pref Nominal Amount at EoY'] = p_eoysec.get('Total Nominal Amount')
    BF['Break Up PC - Total Pref PC Amount at EoY'] = p_eoysec.get('Total Paid-up amount')
    BF['Break Up PC - Total Pref Premium at EoY'] = p_eoysec.get('Total premium')

    return BF


def map_equity_overall(equity_data, BF):
    esc = equity_data.get("Equity Share Capital", {})

    overall = esc.get("Overall", [])
    num_classes = esc.get("Number of Classes")

    # Extract rows safely
    number_row = overall[0] if len(overall) > 0 else {}
    value_row  = overall[1] if len(overall) > 1 else {}
    # === Map Number of Equity Shares ===
    BF["TotalEquitySharesAuthorisationCapital"] = number_row.get("Authorised")
    BF["TotalEquitySharesIssuedCapital"] = number_row.get("Issued")
    BF["TotalEquitySharesSubscribedCapital"] = number_row.get("Subscribed")
    BF["TotalEquitySharesPaidUpCapital"] = number_row.get("Paid Up")

    # === Map Value of Equity Shares (Amount in Rupees) ===
    BF["TotalEquitySharesAuthorisationCapitalValue"] = value_row.get("Authorised")
    BF["TotalEquitySharesIssuedCapitalValue"] = value_row.get("Issued")
    BF["TotalEquitySharesSubscribedCapitalValue"] = value_row.get("Subscribed")
    BF["TotalEquitySharesPaidUpCapitalValue"] = value_row.get("Paid Up")

    # === Number of Equity Share Classes ===
    BF["Number of classesEquity"] = num_classes

    return BF


def map_preference_overall(pref_data, BF):
    psc = pref_data.get("Preference Share Capital", {})

    overall = psc.get("Overall", [])
    num_classes = psc.get("Number of Classes")

    # Extract rows
    number_row = overall[0] if len(overall) > 0 else {}
    value_row  = overall[1] if len(overall) > 1 else {}

    # === Map Number of Preference Shares ===
    BF["TotalPreferenceSharesAuthorisationCapital"] = number_row.get("Authorised")
    BF["TotalPreferenceSharesIssuedCapital"] = number_row.get("Issued")
    BF["TotalPreferenceSharesSubscribedCapital"] = number_row.get("Subscribed")
    BF["TotalPreferenceSharesPaidUpCapital"] = number_row.get("Paid Up")

    # === Map Value of Preference Shares (Amount in Rupees) ===
    BF["TotalPreferenceSharesAuthorisationCapitalValue"] = value_row.get("Authorised")
    BF["TotalPreferenceSharesIssuedCapitalValue"] = value_row.get("Issued")
    BF["TotalPreferenceSharesSubscribedCapitalValue"] = value_row.get("Subscribed")
    BF["TotalPreferenceSharesPaidUpCapitalValue"] = value_row.get("Paid Up")

    # === Number of Preference Share Classes ===
    BF["Number of classes Preference"] = num_classes

    return BF

def map_unclassified_share_capital(data, BF):
    un = data.get("Unclassified Share Capital", {})

    BF["Total Value Unclassified Shares"] = un.get("Total amount of unclassified shares (in rupees)")

    return BF


def map_debenture_capital(debenture_data, BF):
    # Extract each type safely
    ncd = debenture_data.get('Non-convertible debentures', {})
    pcd = debenture_data.get('Partly convertible debentures', {})
    fcd = debenture_data.get('Fully convertible debentures', {})

    # === Non-convertible Debentures (NCD) ===
    BF["NCD BoY"] = ncd.get("Opening")
    BF["NCD Inc DY"] = ncd.get("Increase")
    BF["NCD Dec DY"] = ncd.get("Decrease")
    BF["NCD EoY"] = ncd.get("Closing")

    # === Partly Convertible Debentures (PCD) ===
    BF["PCD BoY"] = pcd.get("Opening")
    BF["PCD Inc DY"] = pcd.get("Increase")
    BF["PCD Dec DY"] = pcd.get("Decrease")
    BF["PCD EoY"] = pcd.get("Closing")

    # === Fully Convertible Debentures (FCD) ===
    BF["FCD BoY"] = fcd.get("Opening")
    BF["FCD Inc DY"] = fcd.get("Increase")
    BF["FCD Dec DY"] = fcd.get("Decrease")
    BF["FCD EoY"] = fcd.get("Closing")

    return BF


def map_non_convertible_debentures_nominal(data, BF):
    ncd_data = data.get("Non-Convertible Debentures", {})
    nominal_table = ncd_data.get("Nominal Value Table", [])

    # Get first (main) entry safely
    row = nominal_table[-1] if len(nominal_table) > 0 else {}

    # === Map fields ===
    BF["NoOfNonConvertibleDebentures"] = row.get("Number of Units")
    BF["AmtPerNonConvertibleDebenture"] = row.get("Nominal Value per Unit")
    BF["Total Value NCD"] = row.get("Total Value (End of Year)")

    return BF


def map_partly_convertible_debentures_nominal(data, BF):
    pcd_data = data.get("Partly Convertible Debentures", {})
    nominal_table = pcd_data.get("Nominal Value Table", [])

    # Get first record safely
    row = nominal_table[-1] if len(nominal_table) > 0 else {}

    # === Map fields ===
    BF["NoOfPartlyConvertibleDebenture"] = row.get("Number of Units")
    BF["AmtPerPartlyConvertibleDebenture"] = row.get("Nominal Value per Unit")
    BF["Total Value PCD"] = row.get("Total Value (End of Year)")

    return BF



def map_fully_convertible_debentures_nominal(data, BF):
    fcd_data = data.get("Fully Convertible Debentures", {})
    nominal_table = fcd_data.get("Nominal Value Table", [])

    # Get first (main) entry safely
    row = nominal_table[-1] if len(nominal_table) > 0 else {}

    # === Map fields ===
    BF["NoOfFullyConvertibleDebenture"] = row.get("Number of Units")
    BF["AmtPerFullyConvertibleDebenture"] = row.get("Nominal Value per Unit")
    BF["Total Value FCD"] = row.get("Total Value (End of Year)")

    return BF



def map_promoter_shareholding(data, BF):
    promoters = data.get('A Promoters', [])

    # Helper: find category safely (case-insensitive contains match)
    def get_cat(keyword):
        for item in promoters:
            if keyword.lower() in item.get('Category', '').lower():
                return item
        return {}

    # === Get Each Category ===
    indian = get_cat('Indian')
    nri = get_cat('NRI')
    foreign_nat = get_cat('Foreign national')
    central_gov = get_cat('Central Government')
    state_gov = get_cat('State Government')
    gov_comp = get_cat('Government companies')
    insurance = get_cat('Insurance')
    banks = get_cat('Banks')
    fin_inst = get_cat('Financial institutions')
    foreign_inst = get_cat('Foreign institutional investors')
    mutual_fund = get_cat('Mutual funds')
    venture_cap = get_cat('Venture capital')
    body_corp = get_cat('Body corporate')
    others = get_cat('Others')
    total = get_cat('Total')

    # === Fill Data into BF ===

    # Indian
    BF["Promoter Eq NoS Indian"] = indian.get('Equity', {}).get('Number of shares')
    BF["Promoter Eq % Indian"] = indian.get('Equity', {}).get('Percentage')
    BF["Promoter Pref NoS Indian"] = indian.get('Preference', {}).get('Number of shares')
    BF["Promoter Pref % Indian"] = indian.get('Preference', {}).get('Percentage')

    # NRI
    BF["Promoter Eq NoS NRI"] = nri.get('Equity', {}).get('Number of shares')
    BF["Promoter Eq % NRI"] = nri.get('Equity', {}).get('Percentage')
    BF["Promoter Pref NoS NRI"] = nri.get('Preference', {}).get('Number of shares')
    BF["Promoter Pref % NRI"] = nri.get('Preference', {}).get('Percentage')

    # Foreign National
    BF["Promoter Eq NoS Foreign National"] = foreign_nat.get('Equity', {}).get('Number of shares')
    BF["Promoter Eq % Foreign National"] = foreign_nat.get('Equity', {}).get('Percentage')
    BF["Promoter Pref NoS Foreign National"] = foreign_nat.get('Preference', {}).get('Number of shares')
    BF["Promoter Pref % Foreign National"] = foreign_nat.get('Preference', {}).get('Percentage')

    # Central Government
    BF["Promoter Eq NoS Central Gov."] = central_gov.get('Equity', {}).get('Number of shares')
    BF["Promoter Eq % Central Gov."] = central_gov.get('Equity', {}).get('Percentage')
    BF["Promoter Pref NoS Central Gov."] = central_gov.get('Preference', {}).get('Number of shares')
    BF["Promoter Pref % Central Gov."] = central_gov.get('Preference', {}).get('Percentage')

    # State Government
    BF["Promoter Eq NoS State Gov."] = state_gov.get('Equity', {}).get('Number of shares')
    BF["Promoter Eq % State Gov."] = state_gov.get('Equity', {}).get('Percentage')
    BF["Promoter Pref NoS State Gov."] = state_gov.get('Preference', {}).get('Number of shares')
    BF["Promoter Pref % State Gov."] = state_gov.get('Preference', {}).get('Percentage')

    # Government Companies
    BF["Promoter Eq NoS Gov Comp."] = gov_comp.get('Equity', {}).get('Number of shares')
    BF["Promoter Eq % Gov Comp."] = gov_comp.get('Equity', {}).get('Percentage')
    BF["Promoter Pref NoS Gov Comp."] = gov_comp.get('Preference', {}).get('Number of shares')
    BF["Promoter Pref % Gov Comp."] = gov_comp.get('Preference', {}).get('Percentage')

    # Insurance
    BF["Promoter Eq NoS Insurance"] = insurance.get('Equity', {}).get('Number of shares')
    BF["Promoter Eq %Insurance"] = insurance.get('Equity', {}).get('Percentage')
    BF["Promoter Pref NoS Insurance"] = insurance.get('Preference', {}).get('Number of shares')
    BF["Promoter Pref %Insurance"] = insurance.get('Preference', {}).get('Percentage')

    # Banks
    BF["Promoter Eq NoS Banks"] = banks.get('Equity', {}).get('Number of shares')
    BF["Promoter Eq % Banks"] = banks.get('Equity', {}).get('Percentage')
    BF["Promoter Pref NoS Banks"] = banks.get('Preference', {}).get('Number of shares')
    BF["Promoter Pref % Banks"] = banks.get('Preference', {}).get('Percentage')

    # Financial Institutions
    BF["Promoter Eq NoS Fin Inst."] = fin_inst.get('Equity', {}).get('Number of shares')
    BF["Promoter Eq % Fin Inst."] = fin_inst.get('Equity', {}).get('Percentage')
    BF["Promoter Pref NoS Fin Inst."] = fin_inst.get('Preference', {}).get('Number of shares')
    BF["Promoter Pref % Fin Inst."] = fin_inst.get('Preference', {}).get('Percentage')

    # Foreign Institutional Investors
    BF["Promoter Eq NoS Foreign Inst."] = foreign_inst.get('Equity', {}).get('Number of shares')
    BF["Promoter Eq % Foreign Inst."] = foreign_inst.get('Equity', {}).get('Percentage')
    BF["Promoter Pref NoS Foreign Inst."] = foreign_inst.get('Preference', {}).get('Number of shares')
    BF["Promoter Pref % Foreign Inst."] = foreign_inst.get('Preference', {}).get('Percentage')

    # Mutual Funds
    BF["Promoter Eq NoS Mutual Funds"] = mutual_fund.get('Equity', {}).get('Number of shares')
    BF["Promoter Eq % Mutual Funds"] = mutual_fund.get('Equity', {}).get('Percentage')
    BF["Promoter Pref NoS Mutual Funds"] = mutual_fund.get('Preference', {}).get('Number of shares')
    BF["Promoter Pref % Mutual Funds"] = mutual_fund.get('Preference', {}).get('Percentage')

    # Venture Capital
    BF["Promoter Eq NoS Venture Capital"] = venture_cap.get('Equity', {}).get('Number of shares')
    BF["Promoter Eq % Venture Capital"] = venture_cap.get('Equity', {}).get('Percentage')
    BF["Promoter Pref NoS Venture Capital"] = venture_cap.get('Preference', {}).get('Number of shares')
    BF["Promoter Pref % Venture Capital"] = venture_cap.get('Preference', {}).get('Percentage')

    # Body Corporate
    BF["Promoter Eq NoS Body Corporate"] = body_corp.get('Equity', {}).get('Number of shares')
    BF["Promoter Eq % Body Corporate"] = body_corp.get('Equity', {}).get('Percentage')
    BF["Promoter Pref NoS Body Corporate"] = body_corp.get('Preference', {}).get('Number of shares')
    BF["Promoter Pref % Body Corporate"] = body_corp.get('Preference', {}).get('Percentage')

    # Others
    BF["Promoter Eq NoS Others"] = others.get('Equity', {}).get('Number of shares')
    BF["Promoter Eq % Others"] = others.get('Equity', {}).get('Percentage')
    BF["Promoter Pref NoS Others"] = others.get('Preference', {}).get('Number of shares')
    BF["Promoter Pref % Others"] = others.get('Preference', {}).get('Percentage')

    # Total
    BF["Promoter Eq NoS Total"] = total.get('Equity', {}).get('Number of shares')
    BF["Promoter Eq % Total"] = total.get('Equity', {}).get('Percentage')
    BF["Promoter Pref NoS Total"] = total.get('Preference', {}).get('Number of shares')
    BF["Promoter Pref % Total"] = total.get('Preference', {}).get('Percentage')

    return BF


def map_public_shareholding(data, BF):
    public = data.get('B Other than promoters', [])

    # Helper: find category safely (case-insensitive contains match)
    def get_cat(keyword):
        for item in public:
            if keyword.lower() in item.get('Category', '').lower():
                return item
        return {}

    # === Get Each Category ===
    indian = get_cat('Indian')
    nri = get_cat('NRI')
    foreign_nat = get_cat('Foreign national')
    central_gov = get_cat('Central Government')
    state_gov = get_cat('State Government')
    gov_comp = get_cat('Government companies')
    insurance = get_cat('Insurance')
    banks = get_cat('Banks')
    fin_inst = get_cat('Financial institutions')
    foreign_inst = get_cat('Foreign institutional investors')
    mutual_fund = get_cat('Mutual funds')
    venture_cap = get_cat('Venture capital')
    body_corp = get_cat('Body corporate')
    others = get_cat('Others')
    total = get_cat('Total')

    # === Fill Data into BF ===

    # Indian
    BF["Pub Promoter Eq NoS Indian"] = indian.get('Equity', {}).get('Number of shares')
    BF["Pub Promoter Eq % Indian"] = indian.get('Equity', {}).get('Percentage')
    BF["Pub Promoter Pref NoS Indian"] = indian.get('Preference', {}).get('Number of shares')
    BF["Pub Promoter Pref % Indian"] = indian.get('Preference', {}).get('Percentage')

    # NRI
    BF["Pub Promoter Eq NoS NRI"] = nri.get('Equity', {}).get('Number of shares')
    BF["Pub Promoter Eq % NRI"] = nri.get('Equity', {}).get('Percentage')
    BF["Pub Promoter Pref NoS NRI"] = nri.get('Preference', {}).get('Number of shares')
    BF["Pub Promoter Pref % NRI"] = nri.get('Preference', {}).get('Percentage')

    # Foreign National
    BF["Pub Promoter Eq NoS Foreign National"] = foreign_nat.get('Equity', {}).get('Number of shares')
    BF["Pub Promoter Eq % Foreign National"] = foreign_nat.get('Equity', {}).get('Percentage')
    BF["Pub Promoter Pref NoS Foreign National"] = foreign_nat.get('Preference', {}).get('Number of shares')
    BF["Pub Promoter Pref % Foreign National"] = foreign_nat.get('Preference', {}).get('Percentage')

    # Central Government
    BF["Pub Promoter Eq NoS Central Gov."] = central_gov.get('Equity', {}).get('Number of shares')
    BF["Pub Promoter Eq % Central Gov."] = central_gov.get('Equity', {}).get('Percentage')
    BF["Pub Promoter Pref NoS Central Gov."] = central_gov.get('Preference', {}).get('Number of shares')
    BF["Pub Promoter Pref % Central Gov."] = central_gov.get('Preference', {}).get('Percentage')

    # State Government
    BF["Pub Promoter Eq NoS State Gov."] = state_gov.get('Equity', {}).get('Number of shares')
    BF["Pub Promoter Eq % State Gov."] = state_gov.get('Equity', {}).get('Percentage')
    BF["Pub Promoter Pref NoS State Gov."] = state_gov.get('Preference', {}).get('Number of shares')
    BF["Pub Promoter Pref % State Gov."] = state_gov.get('Preference', {}).get('Percentage')

    # Government Companies
    BF["Pub Promoter Eq NoS Gov Comp."] = gov_comp.get('Equity', {}).get('Number of shares')
    BF["Pub Promoter Eq % Gov Comp."] = gov_comp.get('Equity', {}).get('Percentage')
    BF["Pub Promoter Pref NoS Gov Comp."] = gov_comp.get('Preference', {}).get('Number of shares')
    BF["Pub Promoter Pref % Gov Comp."] = gov_comp.get('Preference', {}).get('Percentage')

    # Insurance
    BF["Pub Promoter Eq NoS Insurance"] = insurance.get('Equity', {}).get('Number of shares')
    BF["Pub Promoter Eq %Insurance"] = insurance.get('Equity', {}).get('Percentage')
    BF["Pub Promoter Pref NoS Insurance"] = insurance.get('Preference', {}).get('Number of shares')
    BF["Pub Promoter Pref %Insurance"] = insurance.get('Preference', {}).get('Percentage')

    # Banks
    BF["Pub Promoter Eq NoS Banks"] = banks.get('Equity', {}).get('Number of shares')
    BF["Pub Promoter Eq % Banks"] = banks.get('Equity', {}).get('Percentage')
    BF["Pub Promoter Pref NoS Banks"] = banks.get('Preference', {}).get('Number of shares')
    BF["Pub Promoter Pref % Banks"] = banks.get('Preference', {}).get('Percentage')

    # Financial Institutions
    BF["Pub Promoter Eq NoS Fin Inst."] = fin_inst.get('Equity', {}).get('Number of shares')
    BF["Pub Promoter Eq % Fin Inst."] = fin_inst.get('Equity', {}).get('Percentage')
    BF["Pub Promoter Pref NoS Fin Inst."] = fin_inst.get('Preference', {}).get('Number of shares')
    BF["Pub Promoter Pref % Fin Inst."] = fin_inst.get('Preference', {}).get('Percentage')

    # Foreign Institutional Investors
    BF["Pub Promoter Eq NoS Foreign Inst."] = foreign_inst.get('Equity', {}).get('Number of shares')
    BF["Pub Promoter Eq % Foreign Inst."] = foreign_inst.get('Equity', {}).get('Percentage')
    BF["Pub Promoter Pref NoS Foreign Inst."] = foreign_inst.get('Preference', {}).get('Number of shares')
    BF["Pub Promoter Pref % Foreign Inst."] = foreign_inst.get('Preference', {}).get('Percentage')

    # Mutual Funds
    BF["Pub Promoter Eq NoS Mutual Funds"] = mutual_fund.get('Equity', {}).get('Number of shares')
    BF["Pub Promoter Eq % Mutual Funds"] = mutual_fund.get('Equity', {}).get('Percentage')
    BF["Pub Promoter Pref NoS Mutual Funds"] = mutual_fund.get('Preference', {}).get('Number of shares')
    BF["Pub Promoter Pref % Mutual Funds"] = mutual_fund.get('Preference', {}).get('Percentage')

    # Venture Capital
    BF["Pub Promoter Eq NoS Venture Capital"] = venture_cap.get('Equity', {}).get('Number of shares')
    BF["Pub Promoter Eq % Venture Capital"] = venture_cap.get('Equity', {}).get('Percentage')
    BF["Pub Promoter Pref NoS Venture Capital"] = venture_cap.get('Preference', {}).get('Number of shares')
    BF["Pub Promoter Pref % Venture Capital"] = venture_cap.get('Preference', {}).get('Percentage')

    # Body Corporate
    BF["Pub Promoter Eq NoS Body Corporate"] = body_corp.get('Equity', {}).get('Number of shares')
    BF["Pub Promoter Eq % Body Corporate"] = body_corp.get('Equity', {}).get('Percentage')
    BF["Pub Promoter Pref NoS Body Corporate"] = body_corp.get('Preference', {}).get('Number of shares')
    BF["Pub Promoter Pref % Body Corporate"] = body_corp.get('Preference', {}).get('Percentage')

    # Others
    BF["Pub Promoter Eq NoS Others"] = others.get('Equity', {}).get('Number of shares')
    BF["Pub Promoter Eq % Others"] = others.get('Equity', {}).get('Percentage')
    BF["Pub Promoter Pref NoS Others"] = others.get('Preference', {}).get('Number of shares')
    BF["Pub Promoter Pref % Others"] = others.get('Preference', {}).get('Percentage')

    # Total
    BF["Pub Promoter Eq NoS Total"] = total.get('Equity', {}).get('Number of shares')
    BF["Pub Promoter Eq % Total"] = total.get('Equity', {}).get('Percentage')
    BF["Pub Promoter Pref NoS Total"] = total.get('Preference', {}).get('Number of shares')
    BF["Pub Promoter Pref % Total"] = total.get('Preference', {}).get('Percentage')

    return BF

def map_board_of_directors_composition(data, BF):
    board = data.get("A Composition of Board of Directors", [])

    # helper to fetch a category safely by partial match
    def get_cat(keyword):
        for item in board:
            if keyword.lower() in item.get("Category", "").lower():
                return item
        return {}

    # === Get each category ===
    promoter = get_cat("A Promoter")
    non_promoter = get_cat("B Non-Promoter")
    non_independent = get_cat("i Non-Independent")
    independent = get_cat("ii Independent")
    nominee = get_cat("C Nominee Directors")
    nominee_banks = get_cat("i. Banks")
    nominee_invest = get_cat("ii Investing")
    nominee_gov = get_cat("iii Government")
    nominee_small = get_cat("iv Small share")
    nominee_others = get_cat("v Others")
    total = get_cat("Total")

    # === Fill Data ===

    # ---- Executive ----
    BF["No. of Exec Director (Promoter)"] = promoter.get("Number of directors at the end of the year", {}).get("Executive")
    BF["No. of Exec Director (Non -Promoter)"] = non_promoter.get("Number of directors at the end of the year", {}).get("Executive")
    BF["No. of Ind. Exec Director (Non -Promoter)"] = independent.get("Number of directors at the end of the year", {}).get("Executive")
    BF["No. of Non Ind. Exec Director (Non -Promoter)"] = non_independent.get("Number of directors at the end of the year", {}).get("Executive")

    # ---- Nominee Executive ----
    BF["No. of Nominee Exec Director"] = nominee.get("Number of directors at the end of the year", {}).get("Executive")
    BF["No. of Nominee Exec Director (Banks& FII)"] = nominee_banks.get("Number of directors at the end of the year", {}).get("Executive")
    BF["No. of Nominee Exec Director (Investing Insitutions)"] = nominee_invest.get("Number of directors at the end of the year", {}).get("Executive")
    BF["No. of Nominee Exec Director (Govt.)"] = nominee_gov.get("Number of directors at the end of the year", {}).get("Executive")
    BF["No. of Nominee Exec Director (Small Share Holders)"] = nominee_small.get("Number of directors at the end of the year", {}).get("Executive")
    BF["No. of Nominee Exec Director (Others)"] = nominee_others.get("Number of directors at the end of the year", {}).get("Executive")
    BF["No. of Nominee Exec Director (Total)"] = total.get("Number of directors at the end of the year", {}).get("Executive")

    # ---- Non-Executive ----
    BF["No. of Non-Exec Director (Promoter)"] = promoter.get("Number of directors at the end of the year", {}).get("Non-executive")
    BF["No. of Non-Exec Director (Non -Promoter)"] = non_promoter.get("Number of directors at the end of the year", {}).get("Non-executive")
    BF["No. of Ind. Non-Exec Director (Non -Promoter)"] = independent.get("Number of directors at the end of the year", {}).get("Non-executive")
    BF["No. of Non Ind. Non-Exec Director (Non -Promoter)"] = non_independent.get("Number of directors at the end of the year", {}).get("Non-executive")

    # ---- Nominee Non-Executive ----
    BF["No. of Nominee Non-Exec Director"] = nominee.get("Number of directors at the end of the year", {}).get("Non-executive")
    BF["No. of Nominee Non-Exec Director (Banks& FII)"] = nominee_banks.get("Number of directors at the end of the year", {}).get("Non-executive")
    BF["No. of Nominee Non-Exec Director (Investing Insitutions)"] = nominee_invest.get("Number of directors at the end of the year", {}).get("Non-executive")
    BF["No. of Nominee Non-Exec Director (Govt.)"] = nominee_gov.get("Number of directors at the end of the year", {}).get("Non-executive")
    BF["No. of Nominee Non-Exec Director (Small Share Holders)"] = nominee_small.get("Number of directors at the end of the year", {}).get("Non-executive")
    BF["No. of Nominee Non-Exec Director (Others)"] = nominee_others.get("Number of directors at the end of the year", {}).get("Non-executive")
    BF["No. of Nominee Non-Exec Director (Total)"] = total.get("Number of directors at the end of the year", {}).get("Non-executive")

    # ---- % Shares held by Executive Directors ----
    BF["% shares by Exec Director (Promoter)"] = promoter.get("Percentage of shares held by directors as at the end of year", {}).get("Executive")
    BF["% shares by Exec Director (Non -Promoter)"] = non_promoter.get("Percentage of shares held by directors as at the end of year", {}).get("Executive")
    BF["% shares by Ind. Exec Director (Non -Promoter)"] = independent.get("Percentage of shares held by directors as at the end of year", {}).get("Executive")
    BF["% shares by Non Ind. Exec Director (Non -Promoter)"] = non_independent.get("Percentage of shares held by directors as at the end of year", {}).get("Executive")

    # ---- % Shares held by Nominee Executive Directors ----
    BF["% shares by Nominee Exec Director"] = nominee.get("Percentage of shares held by directors as at the end of year", {}).get("Executive")
    BF["% shares by Nominee Exec Director (Banks& FII)"] = nominee_banks.get("Percentage of shares held by directors as at the end of year", {}).get("Executive")
    BF["% shares by Nominee Exec Director (Investing Insitutions)"] = nominee_invest.get("Percentage of shares held by directors as at the end of year", {}).get("Executive")
    BF["% shares by Nominee Exec Director (Govt.)"] = nominee_gov.get("Percentage of shares held by directors as at the end of year", {}).get("Executive")
    BF["% shares by Nominee Exec Director (Small Share Holders)"] = nominee_small.get("Percentage of shares held by directors as at the end of year", {}).get("Executive")
    BF["% shares by Nominee Exec Director (Others)"] = nominee_others.get("Percentage of shares held by directors as at the end of year", {}).get("Executive")
    BF["% shares by Nominee Exec Director (Total)"] = total.get("Percentage of shares held by directors as at the end of year", {}).get("Executive")

    # ---- % Shares held by Non-Executive Directors ----
    BF["% shares by Non-Exec Director (Promoter)"] = promoter.get("Percentage of shares held by directors as at the end of year", {}).get("Non-executive")
    BF["% shares by Non-Exec Director (Non -Promoter)"] = non_promoter.get("Percentage of shares held by directors as at the end of year", {}).get("Non-executive")
    BF["% shares by Ind. Non-Exec Director (Non -Promoter)"] = independent.get("Percentage of shares held by directors as at the end of year", {}).get("Non-executive")
    BF["% shares by Non Ind. Non-Exec Director (Non -Promoter)"] = non_independent.get("Percentage of shares held by directors as at the end of year", {}).get("Non-executive")

    # ---- % Shares held by Nominee Non-Executive Directors ----
    BF["% shares by Nominee Non-Exec Director"] = nominee.get("Percentage of shares held by directors as at the end of year", {}).get("Non-executive")
    BF["% shares by Nominee Non-Exec Director (Banks& FII)"] = nominee_banks.get("Percentage of shares held by directors as at the end of year", {}).get("Non-executive")
    BF["% shares by Nominee Non-Exec Director (Investing Insitutions)"] = nominee_invest.get("Percentage of shares held by directors as at the end of year", {}).get("Non-executive")
    BF["% shares by Nominee Non-Exec Director (Govt.)"] = nominee_gov.get("Percentage of shares held by directors as at the end of year", {}).get("Non-executive")
    BF["% shares by Nominee Non-Exec Director (Small Share Holders)"] = nominee_small.get("Percentage of shares held by directors as at the end of year", {}).get("Non-executive")
    BF["% shares by Nominee Non-Exec Director (Others)"] = nominee_others.get("Percentage of shares held by directors as at the end of year", {}).get("Non-executive")
    BF["% shares by Nominee Non-Exec Director (Total)"] = total.get("Percentage of shares held by directors as at the end of year", {}).get("Non-executive")

    return BF


def map_number_of_promoters_members_debenture(data, BF):
    num_data = data.get('VII NUMBER OF PROMOTERS', [])

    # Helper to find entry by partial match
    def get_cat(keyword):
        for item in num_data:
            for k, v in item.items():
                if keyword.lower() in k.lower():
                    return v
        return {}

    promoters = get_cat("Promoters")
    members = get_cat("Members")
    debentures = get_cat("Debenture")

    # === Fill Data into BF ===
    BF["No. of Promoters BoY"] = promoters.get("At the beginning of the year")
    BF["No. of Promoters EoY"] = promoters.get("At the end of the year")

    BF["No. of Members BoY"] = members.get("At the beginning of the year")
    BF["No. of Members EoY"] = members.get("At the end of the year")

    BF["No. of Debenture Holders BoY"] = debentures.get("At the beginning of the year")
    BF["No. of Debenture Holders EoY"] = debentures.get("At the end of the year")

    return BF


def map_registrar_transfer_agent(data, BF):

    # Directly extract from the given dictionary


    BF["CIN of Registrar/Transfer Agent"] = data.get("CIN of the Registrar and Transfer Agent")
    BF["Name of Registrar/Transfer Agent"] = data.get("Name of the Registrar and Transfer Agent","").title()
    BF["Address of Registrar/Transfer Agent"] = data.get(
        "Registered office address of the Registrar and Transfer Agents"
    )


    return BF



def map_turnover_networth_shareholding(data, BF):
    turnover_data = data.get("V Turnover and Net Worth", {})
    share_data = data.get("VI Share Holding Pattern", {})
    principal_data = data.get("II Principal Business Activities", {})
    BF["No. of Principal Businesses"] = principal_data.get("Number of business activities")
    # === Turnover and Net Worth ===
    BF["Turnover"] = turnover_data.get("Turnover")
    BF["Net Worth"] = turnover_data.get("Net worth of the Company")

    # === Shareholding Pattern ===
    BF["No. of Promoters"] = share_data.get("Total number of shareholders (promoters)")
    BF["Total Number of Shareholders Non Promoters"] = share_data.get("Total number of shareholders (other than promoters)")
    BF["Total Shareholders"] = share_data.get("Total number of shareholders (Promoters + Public/Other than promoters)")

    return BF


#=======================finacial details ========================================================



def process_basic_finacial_pdf(pdf_path):
    # print(f"Processing: {pdf_path}")

    BF = Basic_Finacials.get_basic_financials_headers()

    # COMPANY INFO
    company_raw = company_info.extract_company_info(pdf_path)
    update_company_info(company_raw['company_info'], BF)

    # BREAK-UP OF PAID-UP SHARE CAPITAL
    if "MGT7A" in pdf_path.upper() or "MGT-7A" in pdf_path.upper():
        break_up_raw = Break_up_of_paid_up_share_capital_Mgt7A.process_mgt7A(pdf_path)
    else:
        break_up_raw = Break_up_of_paid_up_share_capital.process_mgt7(pdf_path)

    update_break_up_of_paid_up_share_capital(break_up_raw[0], BF)

    # EQUITY SHARE CAPITAL
    Equity_share_capital_raw = Equity_share_capital.extract_equity_share_capital(pdf_path)
    map_equity_overall(Equity_share_capital_raw, BF)

    # PREFERENCE SHARE CAPITAL
    Preference_share_capital_raw = Preference_share_capital.extract_preference_share_capital(pdf_path)
    map_preference_overall(Preference_share_capital_raw, BF)

    # UNCLASSIFIED SHARE CAPITAL
    Unclassified_share_capital_raw = Unclassified_share_capital.extract_unclassified_share_capital(pdf_path)
    map_unclassified_share_capital(Unclassified_share_capital_raw, BF)
    print("map_unclassified_share_capital")
    # Summary of Indebtedness ==================

    if "MGT7A" in pdf_path.upper() or "MGT-7A" in pdf_path.upper():
        print("Summary_of_Indebtedness table not found ")
    else:
        Summary_of_Indebtedness_raw = Summary_of_Indebtedness.extract_indebtedness_summary(pdf_path)
        map_debenture_capital(Summary_of_Indebtedness_raw,BF)


    # non_convertible_debentures_nominal ======================
    non_convertible_debentures_nominal_raw = Non_convertible_debentures.extract_non_convertible_debentures(pdf_path)
    map_non_convertible_debentures_nominal(non_convertible_debentures_nominal_raw,BF)

    # =======Partly convertible debentures ==============
    Partly_convertible_debentures_raw = Partly_convertible_debentures.extract_partly_convertible_debentures(pdf_path)
    map_partly_convertible_debentures_nominal(Partly_convertible_debentures_raw,BF)

   # ====== Fully convertible debentures ====================
    Fully_convertible_debentures_raw = Fully_convertible_debentures.extract_fully_convertible_debentures(pdf_path)
    map_fully_convertible_debentures_nominal(Fully_convertible_debentures_raw,BF)


   # ======== promoter shereholder ==================
    promoter_raw = A_Promoters.extract_a_promoters(pdf_path)
    map_promoter_shareholding(promoter_raw, BF)

    # ======== public_shareholding ==================
    public_raw = Public_Other_than_promoters.extract_b_promoters(pdf_path)
    map_public_shareholding(public_raw, BF)

    # ======== BoardComposition ==================
    if "MGT7A" in pdf_path.upper() or "MGT-7A" in pdf_path.upper():
        print("Board_of_Directors table not found ")
    else:
        board_data_raw = Board_of_Directors.process_board_of_directors(pdf_path)
        map_board_of_directors_composition(board_data_raw, BF)

    # ======== BoardComposition ==================
    number_raw = NUMBER_OF_PROMOTERS_MEMBERS_DEBENTURE_HOLDERS.vii_number_of_promoters_member_debenture(pdf_path)
    map_number_of_promoters_members_debenture(number_raw, BF)

    # ======== FinancialSummary ==================
    turnover_share_raw = FinancialSummary.extract_mgt7_with_pdfplumber(pdf_path)
    map_turnover_networth_shareholding(turnover_share_raw, BF)

    # ======== FinancialSummary ==================
    if "MGT7A" in pdf_path.upper() or "MGT-7A" in pdf_path.upper():
        print("Number_of_Registrar_and_Transfer_Agent table not found")

    else:
        rta_raw = Number_of_Registrar_and_Transfer_Agent.extract_clean_registrar_info(pdf_path)
        map_registrar_transfer_agent(rta_raw, BF)

    return BF ,Equity_share_capital_raw,Preference_share_capital_raw



#==================== asosiate comanies_information ===================================
import os
import pandas as pd
# import HOLDING_SUBSIDIARY_ASSOCIATE_COMPANIES
#
# ===== Extract Associate Company Information =====
def Associate_Companies_Information_Tab_items(pdf_path, BF):
    data = HOLDING_SUBSIDIARY_ASSOCIATE_COMPANIES.extract_specific_table(pdf_path)
    associate_rows = []

    for row in data:
        if row.get("relation_type", "").strip().lower() == "associate":
            assoc_dict = {
                "CIN": BF.get("CIN"),
                "Company Name": BF.get("Company Name"),
                "Year": BF.get("Year"),
                "Associated Company Name": row.get("company_name","").title(),
                "Associated Company CIN": row.get("cin"),
                "Associated Company Type": row.get("relation_type"),
                "Associated Pecentage Held": row.get("shareholding_percent"),
            }
            associate_rows.append(assoc_dict)

    if not associate_rows:
        return  [{
            "CIN": BF.get("CIN"),
            "Company Name": BF.get("Company Name"),
            "Year": BF.get("Year"),
            "Associated Company Name": "",
            "Associated Company CIN": "",
            "Associated Company Type": "",
            "Associated Pecentage Held": "",
        }]

    return associate_rows


#=========map_directors_kmp ===========================
# === IMPORTS ===
import Directors_and_KMP_Details_Tab_itemes
import Change_in_Directors_Tab_itemes
import HOLDING_SUBSIDIARY_ASSOCIATE_COMPANIES


# ===================================================================
# ========== MAP DIRECTORS & KMP ====================================
# ===================================================================
def map_directors_kmp(BF, directors_list):
    cin = BF.get("CIN", "")
    company_name = BF.get("Company Name", "")
    year = BF.get("Year", "")

    final_rows = []

    for row in directors_list:
        final_rows.append({
            "CIN": cin,
            "Company Name": company_name,
            "Year": year,
            "DIN Or PAN": row.get("DIN/PAN", ""),
            "Name": row.get("Name", "").lstrip(". ").strip().title(),
            "Designation": row.get("Designation", ""),
            "Shares held": row.get("Shares", "")
        })

    if not final_rows:
        return [{
            "CIN": cin,
            "Company Name": company_name,
            "Year": year,
            "DIN Or PAN": "",
            "Name": "",
            "Designation": "",
            "Shares held": ""
        }

        ]

    return final_rows


# ===================================================================
# ========== MAP CHANGE IN DIRECTORS ================================
# ===================================================================
def map_change_in_directors(BF,change_list):
    cin = BF.get("CIN", "")
    company_name = BF.get("Company Name", "")
    year = BF.get("Year", "")

    final_rows = []

    for row in change_list:
        final_rows.append({
            "CIN": cin,
            "Company Name": company_name,
            "Year": year,
            "DIN Or PAN": row.get("DIN/PAN", ""),
            "Name": row.get("Name", "").lstrip(". ").strip().title(),
            "Designation": row.get("Designation at the beginning / during the financial year", ""),
            "Appointment Date": row.get("Date of appointment/ change in designation/ cessation (DD/MM/YYYY)", ""),
            "Nature of Change": row.get(
                "Nature of change (Appointment/ Change in designation/ Cessation)", ""
            )
        })

    if not final_rows:
        return [{
            "CIN": cin,
            "Company Name": company_name,
            "Year": year,
            "DIN Or PAN": "",
            "Name":"",
            "Designation": "",
            "Appointment Date": "",
            "Nature of Change": ""
        }

        ]

    return final_rows



# ================= Equity share capital  ===========

def map_equity_share_capital(BF, equity_raw):
    cin = BF.get("CIN", "")
    company = BF.get("Company Name", "")
    year = BF.get("Year", "")

    data = equity_raw.get("Equity Share Capital", {})
    classes = data.get("Classes", [])

    # Extract first class row – always contains the Numbers
    class_row = classes[0] if len(classes) > 0 else {}

    # Extract nominal value row
    nominal_row = classes[1] if len(classes) > 1 else {}

    # Extract total value row
    value_row = classes[2] if len(classes) > 2 else {}

    return [{
        "CIN": cin,
        "Company Name": company,
        "Year": year,
        "ClassOfShares": "",#class_row.get("Class", ""),

        "Authorised Capital Equity Share Number": class_row.get("Authorised", ""),
        "Authorised Capital Equity Share Nominal Value": nominal_row.get("Authorised", ""),
        "Authorised Capital Equity Share Value": value_row.get("Authorised", ""),

        "Issued Capital Equity Share Number": class_row.get("Issued", ""),
        "Issued Capital Equity Share Nominal Value": nominal_row.get("Issued", ""),
        "Issued Capital Equity Share Value": value_row.get("Issued", ""),

        "Subscribed Capital Equity Share Number": class_row.get("Subscribed", ""),
        "Subscribed Capital Equity Share Nominal Value": nominal_row.get("Subscribed", ""),
        "Subscribed Capital Equity Share Value": value_row.get("Subscribed", ""),

        "Paidup Capital Equity Share Number": class_row.get("Paid Up", ""),
        "Paidup Capital Equity Share Nominal Value": nominal_row.get("Paid Up", ""),
        "Paidup Capital Equity Share Value": value_row.get("Paid Up", "")
    }]


# ==================== prefrence share capital ==================================

def map_preference_share_capital(BF, pref_raw):
    cin = BF.get("CIN", "")
    company = BF.get("Company Name", "")
    year = BF.get("Year", "")

    data = pref_raw.get("Preference Share Capital", {})
    classes = data.get("Classes", [])

    class_row = classes[0] if len(classes) > 0 else {}
    nominal_row = classes[1] if len(classes) > 1 else {}
    value_row = classes[2] if len(classes) > 2 else {}

    return [{
        "CIN": cin,
        "Company Name": company,
        "Year": year,

        "Authorised Capital Preference Share Number": class_row.get("Authorised", ""),
        "Authorised Capital Preference Share Nominal Value": nominal_row.get("Authorised", ""),
        "Authorised Capital Preference Share Value": value_row.get("Authorised", ""),

        "Issued Capital Preference Share Number": class_row.get("Issued", ""),
        "Issued Capital Preference Share Nominal Value": nominal_row.get("Issued", ""),
        "Issued Capital Preference Share Value": value_row.get("Issued", ""),

        "Subscribed Capital Preference Share Number": class_row.get("Subscribed", ""),
        "Subscribed Capital Preference Share Nominal Value": nominal_row.get("Subscribed", ""),
        "Subscribed Capital Preference Share Value": value_row.get("Subscribed", ""),

        "Paidup Capital Preference Share Number": class_row.get("Paid Up", ""),
        "Paidup Capital Preference Share Nominal Value": nominal_row.get("Paid Up", ""),
        "Paidup Capital Preference Share Value": value_row.get("Paid Up", "")
    }]

# ===================================================================
# ========== SAVE MULTIPLE SHEETS TO EXCEL ==========================
# ===================================================================
import X_REMUNERATION_OF_DIRECTORS_AND_KEY_MANAGERIAL_PERSONNEL
import os
import pandas as pd
import openpyxl

import PRINCIPAL_BUSINESS_ACTIVITIES
# ============================================================
# REMUNERATION HEADERS
# ============================================================
REMUN_HEADERS = [
    "CIN", "Company Name", "Year",
    "Name", "Designation",
    "Gross Salary", "Commission", "Stocks", "Others", "Total"
]



def remove_total_row(rows):
    cleaned = []
    for r in rows:
        if str(r.get("Name", "")).strip().lower() == "total":
            break
        cleaned.append(r)
    return cleaned



# ============================================================
# MAP SINGLE REMUNERATION TABLE
# ============================================================
def map_remuneration_table(BF, rows):

    cin = BF.get("CIN", "")
    comp = BF.get("Company Name", "")
    year = BF.get("Year", "")

    mapped = []

    for r in rows or []:
        mapped.append({
            "CIN": cin,
            "Company Name": comp,
            "Year": year,
            "Name": r.get("Name", ""),
            "Designation": r.get("Designation", ""),
            "Gross Salary": r.get("Gross salary", ""),
            "Commission": r.get("Commission", ""),
            "Stocks": r.get("Stock Option/ Sweat equity", ""),
            "Others": r.get("Others", ""),
            "Total": r.get("Total amount", "")
        })

    if not mapped:
       return [ {
            "CIN": cin,
            "Company Name": comp,
            "Year": year,
            "Name": "",
            "Designation": "",
            "Gross Salary": "",
            "Commission": "",
            "Stocks": "",
            "Others": "",
            "Total": ""
        }]

    # ✅ ALWAYS return list
    return mapped


# ============================================================
# SPLIT 3 TABLES — ONLY IF ALL THREE HAVE DATA
# ============================================================
def split_three_remuneration_tables_mgt7(BF, raw):

    t1 = remove_total_row(raw.get("table_1", []))
    t2 = remove_total_row(raw.get("table_2", []))
    t3 = remove_total_row(raw.get("table_3", []))

    # All 3 must have data — otherwise skip all
    if not t1 or not t2 or not t3:
        return None, None, None

    return (
        map_remuneration_table(BF, t1),
        map_remuneration_table(BF, t2),
        map_remuneration_table(BF, t3)
    )

def split_three_remuneration_tables_mgt7A(BF, raw):

    t1 = remove_total_row(raw.get("table_1", []))
    # t2 = remove_total_row([])
    t3 = remove_total_row(raw.get("table_2", []))

    # All 3 must have data — otherwise skip all
    if not t1 or not t3:
        return None, None, None

    return (
        map_remuneration_table(BF, t1),
        map_remuneration_table(BF, []),
        map_remuneration_table(BF, t3)
    )


def PRINCIPAL_BUSINESS_ACTIVITIES_map(BF, business_list):
    """
    Safe mapping function → Works even if business_list is None or not a list.
    """

    # Ensure always a list
    if not isinstance(business_list, list):
        business_list = []

    result = {
        "CIN": BF.get("CIN", ""),
        "Year": BF.get("Year", ""),
        "Pan of Company":"", # BF.get("Pan of Company", ""),
        "Company Name": BF.get("Company Name", "")
    }

    # Loop through each business row (max 10)
    for idx, row in enumerate(business_list[:10], start=1):

        result[f"Principal Business Description {idx}"] = row.get(
            "Description of Business Activity", ""
        )

        result[f"Principal Business Activity {idx}"] = row.get(
            "Business Activity Code", ""
        )

        result[f"Percentage Contribution in Turnover {idx}"] = row.get(
            "% of turnover of the company", ""
        )

    # Ensure all 10 fields always exist
    total = len(business_list)
    for idx in range(total + 1, 11):
        result[f"Principal Business Description {idx}"] = ""
        result[f"Principal Business Activity {idx}"] = ""
        result[f"Percentage Contribution in Turnover {idx}"] = ""

    return result


def map_company_penalty_sheet(BF):
    """
    Creates one empty row for Penalty/Offence sheet
    using only BF fields.
    """

    return [{
        "CIN": BF.get("CIN", ""),
        "Company Name": BF.get("Company Name", ""),
        "Year": BF.get("Year", ""),
        "Company or Party Name": "",
        "Court Name": "",
        "Date of Order": "",
        "Name of Act": "",
        "Details of Offence": "",
        "Status Of Appeal": ""
    }]


def map_compound_penalty_sheet(BF):
    """
    Creates one empty row for Compound Penalty sheet
    using only BF fields.
    """

    return [{
        "CIN": BF.get("CIN", ""),
        "Company Name": BF.get("Company Name", ""),
        "Year": BF.get("Year", ""),
        "Company or Party Name": "",
        "Court Name": "",
        "Date of Order": "",
        "Name of Act": "",
        "Details of Offence": "",
        "Status Of Appeal": "",
        "Compound Amount": ""
    }]

def map_securities_other_than_shares(BF):
    """
    Creates one empty row for
    'Securities Other Than Shares' sheet
    using only BF values.
    """

    return [{
        "CIN": BF.get("CIN", ""),
        "Company Name": BF.get("Company Name", ""),
        "Year": BF.get("Year", ""),
        "Type Of Security": "",
        "Number of Securities": "",
        "Nominal Value Per Security": "",
        "Total Nominal Value": "",
        "Paidup Value per Security": "",
        "Total Paidup Value": ""
    }]
# ============================================================
# SAVE NORMAL SHEETS (NO REMUNERATION)
# ============================================================

def append_df_to_excel(filename, df, sheet_name):
    import openpyxl
    from openpyxl.utils.dataframe import dataframe_to_rows

    # ==== CASE 1: FILE DOES NOT EXIST → CREATE NEW ====
    if not os.path.isfile(filename):
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        return

    # ==== CASE 2: FILE EXISTS → APPEND ====
    wb = openpyxl.load_workbook(filename)

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        # Write header + rows
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
    else:
        ws = wb[sheet_name]
        # Append rows WITHOUT header
        for r in dataframe_to_rows(df, index=False, header=False):
            ws.append(r)

    wb.save(filename)


def save_to_excel(BF, associates, directors, change_directors,
                  equity_data, pref_data,
                  dir_rem, ceo_rem, other_rem,PRINCIPAL_BUSINESS_ACTIVITIES_data,
                  company_penalty,compound_penalty,securities_other_than,
                  output_excel):

    append_df_to_excel(output_excel, pd.DataFrame([BF]), "Basic Financials")
    if PRINCIPAL_BUSINESS_ACTIVITIES_data:
        append_df_to_excel(output_excel, pd.DataFrame([PRINCIPAL_BUSINESS_ACTIVITIES_data]), "Principal Business Activities")
    append_df_to_excel(output_excel, pd.DataFrame(associates), "Associate Companies Information")
    append_df_to_excel(output_excel, pd.DataFrame(directors), "Directors and KMP Details")
    append_df_to_excel(output_excel, pd.DataFrame(change_directors), "Change in Directors")
    append_df_to_excel(output_excel, pd.DataFrame(equity_data), "Equity Shares Details")
    append_df_to_excel(output_excel, pd.DataFrame(pref_data), "Preference Shares Details")
    append_df_to_excel(output_excel, pd.DataFrame(company_penalty), "Company Penalty")
    append_df_to_excel(output_excel, pd.DataFrame(compound_penalty), "Company Compound Penalty")
    append_df_to_excel(output_excel, pd.DataFrame(securities_other_than), "Securities Other than Shares")
    # --- Only when tables exist ---
    if dir_rem:
        append_df_to_excel(output_excel, pd.DataFrame(dir_rem), "Director's Remunerations")

    if ceo_rem:
        append_df_to_excel(output_excel, pd.DataFrame(ceo_rem), "CEO Remunerations")

    if other_rem:
        append_df_to_excel(output_excel, pd.DataFrame(other_rem), "Others Remunerations")


# ============================================================
# MAIN LOOP
# ============================================================
input_folder = r"C:\Users\PC\Downloads\mgt"
output_excel = r"C:\Users\PC\Downloads\mgt\MGT7_And_MGT7A_final_report_2.xlsx"


def main():

    for file in os.listdir(input_folder):
        if not file.lower().endswith(".pdf"):
            continue

        pdf = os.path.join(input_folder, file)
        print(f"Processing → {file}")

        file = file.lower()
        # ✅ Filter only MGT-7 / MGT-7A of 2025
        if not (("mgt7" in file or "mgt-7" in file) and "2025" in file):
            print(f"⏭ Skipped (not 2025 MGT-7): {file}")
            continue
    # for root, dirs, files in os.walk(input_folder):
    #
    #     # :white_check_mark: Process only "Annual" folder
    #     if "annual" not in os.path.basename(root).lower() :
    #         continue
    #     for file in files:
    #         if not file.lower().endswith(".pdf"):
    #             continue
    #         file_lower = file.lower()
    #         # :white_check_mark: Only MGT-7 / MGT-7A of 2025
    #         if not (
    #             ("mgt7_form" in file_lower or "mgt-7_form" in file_lower or "mgt7form" in file_lower or "mgt-7form" in file_lower)
    #             and "2025" in file_lower
    #         ):
    #             print(f":black_right_pointing_double_triangle_with_vertical_bar: Skipped (not 2025 MGT-7): {file}")
    #             continue
    #         pdf = os.path.join(root, file)
        try:
            BF, equity_raw, pref_raw = process_basic_finacial_pdf(pdf)

            associates = Associate_Companies_Information_Tab_items(pdf, BF)

            if "MGT7A" in pdf.upper() or "MGT-7A" in pdf.upper():
                change_directors = map_change_in_directors(BF,[])
                directors = directors = map_directors_kmp(BF,[])
            else:
                change_directors = map_change_in_directors(BF, Change_in_Directors_Tab_itemes.extract_bii(pdf))
                directors = map_directors_kmp(BF, Directors_and_KMP_Details_Tab_itemes.extract_bi_directors_clean_final(pdf))

            print("directors")
            equity_data = map_equity_share_capital(BF, equity_raw)
            pref_data = map_preference_share_capital(BF, pref_raw)
            print("pref_data")
            PRINCIPAL_BUSINESS_ACTIVITIES_raw = PRINCIPAL_BUSINESS_ACTIVITIES.extrect_PRINCIPAL_BUSINESS_ACTIVITIES(pdf)
            PRINCIPAL_BUSINESS_ACTIVITIES_data = PRINCIPAL_BUSINESS_ACTIVITIES_map(BF,PRINCIPAL_BUSINESS_ACTIVITIES_raw)
            # Remuneration extract
            print("PRINCIPAL_BUSINESS_ACTIVITIES_data")

            if "MGT7A" in pdf.upper() or "MGT-7A" in pdf.upper():
                rem_raw = X_REMUNERATION_OF_DIRECTORS_AND_KEY_MANAGERIAL_PERSONNEL.auto_extract_clean_three(pdf)
                dir_rem, ceo_rem, other_rem = split_three_remuneration_tables_mgt7A(BF, rem_raw)
            else:
                rem_raw = X_REMUNERATION_OF_DIRECTORS_AND_KEY_MANAGERIAL_PERSONNEL.auto_extract_clean_three(pdf)
                dir_rem, ceo_rem, other_rem = split_three_remuneration_tables_mgt7(BF, rem_raw)


            company_penalty = map_company_penalty_sheet(BF)
            print(company_penalty)
            compound_penalty = map_compound_penalty_sheet(BF)
            print(compound_penalty)
            securities_other_than = map_securities_other_than_shares(BF)
            print(securities_other_than)
            save_to_excel(
                BF, associates, directors, change_directors,
                equity_data, pref_data,
                dir_rem, ceo_rem, other_rem,PRINCIPAL_BUSINESS_ACTIVITIES_data,
                company_penalty,compound_penalty,securities_other_than,
                output_excel
            )

            print(f"✔ Done: {file}")

        except Exception as e:

            print(f"❌ Error → {file}: {e}")

    print("\n🎯 COMPLETED ALL PDF\n")


if __name__ == "__main__":
    main()


